# -*- coding: utf-8 -*-
# Excel,CSV, JSON 등등 파일에 대한 읽기, 쓰기 라이브러리 
# Data Factory Studio / Object Storage 에 올리는 파일포멧(CSV, JSON)으로 변환하는 기능 관련 모듈 





import os  
import re 
import pprint 
pp = pprint.PrettyPrinter(indent=2)
import json 
from pathlib import WindowsPath

import json
from openpyxl import load_workbook
from datetime import datetime, date


import pandas as pd


from fileio.core import *



################################################################
# 엑셀 파일
################################################################

class ExcelSheet:
    # xslx_file: 엑셀파일의 파일명 포함 전체경로
    # sheet_name: 엑셀파일 원본 시트명
    # alias_name: DFS에 업로드하기 위해 반드시 필요한 영문 시트명. 데이터모델링 시 클래스명으로 동일하게 사용하길 권장

    def __init__(self, xslx_file, sheet_name, alias_name=None):
        self.filepath = xslx_file
        
        # 시트명으로 모델명 설정
        self.name = sheet_name
        if check_has_korean(sheet_name):
            if alias_name is None:
                self.modelName = None
            else:
                if check_has_korean(alias_name):
                    self.modelName = None
                else:
                    self.modelName = alias_name
        else:
            self.name = sheet_name 
            self.modelName = sheet_name 

        if self.modelName is None:
            print("""
            WARNING | 한글 시트명을 반드시 영문으로 바꿔야 하는 이유:
            [1] DFS에 업로드하기 위해서 파일명 포함 전체 경로 문자열이 영문이 아니면 Pipeline을 사용할 수 없다.
            [2] DFS에서 데이터 모델링 시 클래스명은 반드시 영문이어야 한다. 클래스명은 '파이썬 클래스' 작명 규칙에 따라 작성해야한다.
            """)

    # 엑셀시트의 테이블을 데이터프레임으로 읽어온다
    def load_data(self):
        df = pd.read_excel(self.filepath, sheet_name=self.name)
        return df.dropna(axis=0, how='all')

    # DFS Data Modeling이 요구하는 형태로 스키마 파일구조 생성
    def gen_schema(self):
        df = self.load_data()
        df = restruct_dataframe(df)
        schema = generate_schema(df)
        return schema
    
    def parse_data(self, schema_filepath=None):
        df = self.load_data()
        if schema_filepath is None:
            # 자동으로 변환
            df = normalize_dataframe(df)
        else:
            # 사용자 지정 스키마 기준으로 변환
            # 컬럼명만 변환한다
            schema = read_json(schema_filepath)
            # pp.pprint(schema)
            columns = {}
            for col, dic in schema.items():
                columns.update({col: dic['Column Name']})
            # pp.pprint(columns)

            df = restruct_dataframe(df)
            df = df.rename(columns=columns)

        return df



class ExcelFile:

    def __init__(self, xslx_file):
        # 파일경로 + 파일명 + 확장자
        self.filepath = xslx_file

        # 실제 파일 경로
        self.path = os.path.dirname(xslx_file)
        # 실제 파일 (확장자 포함)
        self.file = os.path.basename(xslx_file)
        # DFS / Data Model / Package Name 으로 사용할 변수명 
        root, ext = os.path.splitext(self.file)
        self.filename = root
        self.setup_sheets()

    def setup_sheets(self):
        try:
            workbook = load_workbook(self.filepath, data_only=True)
            self._sheets = []
            for i, sh_name in enumerate(workbook.sheetnames):
                self._sheets.append((i, sh_name, workbook[sh_name]))
            
            # workbook.close()
            self._workbook = workbook 
        except Exception as e:
            print(f"시트 목록 조회 중 오류 발생: {e}")
            return []
        
    def get_sheet(self, sheet_name):
        for i, sh_name, worksheet in self._sheets:
            if sheet_name == sh_name:
                return WorkSheet(worksheet)
            
    def write_jsonfile(self):
        _dir = os.path.join(self.path, self.filename)
        os.makedirs(_dir, exist_ok=True)

        for i, sh_name, worksheet in self._sheets:
            print('\n\n')
            print([i, sh_name, worksheet])
            ws = WorkSheet(worksheet)
            print(ws.name)
            df = ws.parse_asto_df()
            if isinstance(df, pd.DataFrame):
                jsonfile = os.path.join(_dir, f"{i}_{sh_name}.json")
                df.to_json(
                    jsonfile, 
                    orient="records", 
                    force_ascii=False, 
                    date_format='iso',
                    indent=2,
                )


class WorkSheet:

    def __init__(self, worksheet):
        self._worksheet = worksheet 
        self.name = worksheet._WorkbookChild__title

    def parse(self):
        data = []
        for row in self._worksheet.iter_rows(values_only=True):
            is_data = True 
            for elem in row:
                if elem:
                    pass 
                else:
                    is_data = False 
                    break 

            if is_data:
                data.append(row)
        
        if len(data) > 0:
            columns = data.pop(0)
            columns = [c.strip() for c in columns]
            return columns, data 
        else:
            return None, None 
    
    def parse_asto_df(self):
        cols, data = self.parse()
        if cols:
            return pd.DataFrame(data, columns=cols)
    
    

    



    


